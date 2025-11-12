#!/usr/bin/env python3
"""
Face filter FaceNet + style-aware pipeline.

input/
    person_a/
        <source image>  # shortest filename treated as source
        <filtered images>
    person_b/
        ...
output/
    facenet_scores.xlsx  (matrix 시트 + raw 시트)

실행 예시:
    & "$env:USERPROFILE\Desktop\privacy-aware-face-recognition\venv\Scripts\python.exe" `
        "$env:USERPROFILE\Desktop\filter\facenet_pipeline.py"
"""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Sequence, Tuple

import lpips
import pandas as pd
import torch
from PIL import Image
from facenet_pytorch import InceptionResnetV1
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from torchvision import transforms

ALLOWED_EXTS = {".jpg", ".jpeg", ".png", ".bmp", ".webp"}
LPIPS_SIZE = 224
FACENET_SIZE = 160
STYLE_DISTANCE_CAP_DEFAULT = 0.2


@dataclass(frozen=True)
class PersonFolder:
    name: str
    folder: Path
    original: Path


def resolve_dir(path: str | Path, base: Path) -> Path:
    p = Path(path)
    if not p.is_absolute():
        p = (base / p).resolve()
    return p


def list_person_folders(input_dir: Path) -> List[Path]:
    return [p for p in sorted(input_dir.iterdir()) if p.is_dir()]


def pick_original(img_files: Sequence[Path]) -> Path:
    return min(img_files, key=lambda p: (len(p.stem), len(p.name)))


def image_files(folder: Path) -> List[Path]:
    files = []
    for p in sorted(folder.iterdir()):
        if p.is_file() and p.suffix.lower() in ALLOWED_EXTS:
            files.append(p)
    return files


def canonical_filter_name(variant: Path, original: Path) -> str:
    name = variant.stem
    suffixes = [
        f"_{original.stem}",
        f"-{original.stem}",
        f".{original.stem}",
    ]
    lowered = name.lower()
    for suffix in suffixes:
        if lowered.endswith(suffix.lower()):
            name = name[: -len(suffix)]
            break
    return name or variant.stem


def load_facenet_model(device: str) -> InceptionResnetV1:
    return InceptionResnetV1(pretrained="vggface2").eval().to(device)


def load_lpips_model(device: str, net: str) -> lpips.LPIPS:
    model = lpips.LPIPS(net=net).eval().to(device)
    return model


FACENET_TRANSFORM = transforms.Compose(
    [
        transforms.Resize((FACENET_SIZE, FACENET_SIZE)),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5]),
    ]
)

LPIPS_TRANSFORM = transforms.Compose(
    [
        transforms.Resize((LPIPS_SIZE, LPIPS_SIZE)),
        transforms.ToTensor(),
        transforms.Normalize(mean=[0.5, 0.5, 0.5], std=[0.5, 0.5, 0.5]),
    ]
)


def load_image_for_facenet(path: Path) -> torch.Tensor:
    return FACENET_TRANSFORM(Image.open(path).convert("RGB"))


def load_image_for_lpips(path: Path) -> torch.Tensor:
    return LPIPS_TRANSFORM(Image.open(path).convert("RGB"))


def embed_image(path: Path, model: InceptionResnetV1, device: str) -> torch.Tensor:
    tensor = load_image_for_facenet(path).unsqueeze(0).to(device)
    with torch.no_grad():
        emb = model(tensor).squeeze(0).cpu()
    return emb


def cosine_similarity(a: torch.Tensor, b: torch.Tensor) -> float:
    a = a / (a.norm(p=2) + 1e-8)
    b = b / (b.norm(p=2) + 1e-8)
    return float(torch.mul(a, b).sum().item())


def cosine_to_percent(value: float) -> float:
    return (value + 1.0) * 50.0


def compute_style_distance(
    original: Path, variant: Path, model: lpips.LPIPS, device: str
) -> float:
    img_a = load_image_for_lpips(original).unsqueeze(0).to(device)
    img_b = load_image_for_lpips(variant).unsqueeze(0).to(device)
    with torch.no_grad():
        distance = float(model(img_a, img_b).item())
    return distance


def style_distance_to_percent(distance: float, max_distance: float) -> float:
    d = max(0.0, min(distance, max_distance))
    return (1.0 - d / max_distance) * 100.0


def process_folder(
    folder: Path,
    face_model: InceptionResnetV1,
    style_model: lpips.LPIPS,
    device: str,
    style_threshold: float,
    
    
    style_distance_cap: float,
) -> Tuple[PersonFolder, List[Dict[str, float]]]:
    files = image_files(folder)
    if len(files) < 2:
        raise ValueError(
            f"'{folder.name}' needs at least 2 images (source + filtered variants)."
        )

    original = pick_original(files)
    variants = [p for p in files if p != original]

    orig_emb = embed_image(original, face_model, device)

    records: List[Dict[str, float]] = []
    for variant in variants:
        variant_emb = embed_image(variant, face_model, device)
        cos = cosine_similarity(orig_emb, variant_emb)
        face_percent = cosine_to_percent(cos)

        style_distance = compute_style_distance(original, variant, style_model, device)
        style_percent = style_distance_to_percent(style_distance, style_distance_cap)

        if style_distance >= style_threshold:
            face_weight = 0.3
            applied_style_weight = 0.7
            style_used = True
        else:
            face_weight = 1.0
            applied_style_weight = 0.0
            style_used = False

        combined_percent = face_weight * face_percent + applied_style_weight * style_percent

        records.append(
            {
                "filter_name": canonical_filter_name(variant, original),
                "filter_file": variant.name,
                "original_name": original.stem,
                "original_file": original.name,
                "person_folder": folder.name,
                "face_cosine": round(cos, 4),
                "face_percent": round(face_percent, 2),
                "style_distance": round(style_distance, 4),
                "style_percent": round(style_percent, 2),
                "face_weight": round(face_weight, 2),
                "style_weight": round(applied_style_weight, 2),
                "style_used": style_used,
                "final_percent": round(combined_percent, 2),
            }
        )

    if records:
        distances = [rec["style_distance"] for rec in records]
        used = sum(1 for rec in records if rec["style_used"] )
        print(
            f"    style distance range: {min(distances):.3f} - {max(distances):.3f}, "
            f"style weighting applied: {used}/{len(records)}"
        )

    person = PersonFolder(name=folder.name, folder=folder, original=original)
    return person, records


def build_matrix(
    records: Iterable[Dict[str, float]], column_order: Sequence[str]
) -> pd.DataFrame:
    matrix: Dict[str, Dict[str, float]] = defaultdict(dict)
    for rec in records:
        matrix[rec["filter_name"]][rec["original_name"]] = rec["final_percent"]

    df = pd.DataFrame.from_dict(matrix, orient="index")
    if df.empty:
        return df
    ordered_cols = [c for c in column_order if c in df.columns]
    remaining = [c for c in df.columns if c not in ordered_cols]
    df = df[ordered_cols + remaining]
    df.index.name = "filter_name"
    return df.sort_index().round(2)


def read_existing_raw(excel_path: Path) -> pd.DataFrame:
    if not excel_path.exists():
        return pd.DataFrame()
    try:
        return pd.read_excel(excel_path, sheet_name="raw")
    except ValueError:
        return pd.DataFrame()
    except Exception as exc:  # pylint: disable=broad-except
        print(f"[WARN] Failed to read raw sheet '{excel_path}': {exc}")
        return pd.DataFrame()


def merge_with_existing(
    all_records: List[Dict[str, float]], existing_raw: pd.DataFrame
) -> pd.DataFrame:
    df_new = pd.DataFrame(all_records)
    if df_new.empty:
        return existing_raw
    if existing_raw.empty or "original_name" not in existing_raw.columns:
        return df_new.reset_index(drop=True)

    new_originals = set(df_new["original_name"].tolist())
    df_existing = existing_raw[
        ~existing_raw["original_name"].isin(new_originals)
    ].copy()
    if df_existing.empty:
        return df_new.reset_index(drop=True)
    return pd.concat([df_existing, df_new], ignore_index=True)


def load_existing_column_order(excel_path: Path) -> List[str]:
    if not excel_path.exists():
        return []
    try:
        df_matrix = pd.read_excel(excel_path, sheet_name="matrix", index_col=0)
        return df_matrix.columns.tolist()
    except ValueError:
        return []
    except Exception as exc:  # pylint: disable=broad-except
        print(f"[WARN] Failed to read matrix sheet '{excel_path}': {exc}")
        return []


def export_workbook(df_matrix: pd.DataFrame, df_raw: pd.DataFrame, excel_path: Path) -> None:
    ordered_columns = [
        "person_folder",
        "original_name",
        "original_file",
        "filter_name",
        "filter_file",
        "face_cosine",
        "face_percent",
        "style_distance",
        "style_percent",
        "face_weight",
        "style_weight",
        "style_used",
        "final_percent",
    ]
    extra_cols = [c for c in df_raw.columns if c not in ordered_columns]
    df_raw = df_raw[ordered_columns + extra_cols]

    center_alignment = Alignment(horizontal="center", vertical="center")
    red_font = Font(color="FF0000")

    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        df_matrix.to_excel(writer, sheet_name="matrix", float_format="%.2f", na_rep="")
        matrix_sheet = writer.sheets["matrix"]

        index_width = 20
        if not df_matrix.empty:
            index_width = max(index_width, max(len(str(idx)) for idx in df_matrix.index) + 4)
        matrix_sheet.column_dimensions["A"].width = index_width

        for col_idx, column in enumerate(df_matrix.columns, start=2):
            width = max(14, len(str(column)) + 4)
            matrix_sheet.column_dimensions[get_column_letter(col_idx)].width = width

        max_row = matrix_sheet.max_row
        max_col = matrix_sheet.max_column
        for row in matrix_sheet.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=max_col):
            for cell in row:
                cell.alignment = center_alignment
                try:
                    value = float(cell.value) if cell.value is not None else None
                except (TypeError, ValueError):
                    value = None
                if value is not None and value >= 80.0:
                    cell.font = red_font

        df_raw.to_excel(writer, sheet_name="raw", index=False)
        raw_sheet = writer.sheets["raw"]
        for col_idx, column in enumerate(df_raw.columns, start=1):
            width = max(12, len(str(column)) + 4)
            raw_sheet.column_dimensions[get_column_letter(col_idx)].width = width

        percent_columns = [idx for idx, column in enumerate(df_raw.columns, start=1) if "percent" in column.lower()]
        final_percent_idx = None
        for idx, column in enumerate(df_raw.columns, start=1):
            if column.lower() == "final_percent":
                final_percent_idx = idx
                break

        for col_idx in percent_columns:
            for cell in raw_sheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, max_row=raw_sheet.max_row):
                current = cell[0]
                current.alignment = center_alignment
                if final_percent_idx is not None and col_idx == final_percent_idx:
                    try:
                        value = float(current.value) if current.value is not None else None
                    except (TypeError, ValueError):
                        value = None
                    if value is not None and value >= 80.0:
                        current.font = red_font

def cleanup_csv_artifacts(output_dir: Path) -> None:
    patterns = ["facenet_scores_raw.csv", "reference_*.csv"]
    for pattern in patterns:
        for csv_path in output_dir.glob(pattern):
            try:
                csv_path.unlink()
            except OSError as exc:
                print(f"[WARN] Failed to remove legacy file '{csv_path}': {exc}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="FaceNet + style comparison pipeline")
    parser.add_argument(
        "--input-dir",
        default="input",
        help="Folder that contains per-person sub-folders (default: input)",
    )
    parser.add_argument(
        "--output-dir",
        default="output",
        help="Folder to store Excel output (default: output)",
    )
    parser.add_argument(
        "--excel-name",
        default="facenet_scores.xlsx",
        help="Excel file name",
    )
    parser.add_argument(
        "--style-threshold",
        type=float,
        default=0.02,
        help="LPIPS distance threshold to treat as 강한 스타일 변화",
    )
    
    parser.add_argument(
        "--lpips-net",
        default="vgg",
        help="LPIPS backbone (alex/vgg/squeeze)",
    )
    parser.add_argument(
        "--style-distance-cap",
        type=float,
        default=STYLE_DISTANCE_CAP_DEFAULT,
        help="LPIPS 거리를 이 값으로 클램핑 후 퍼센트로 변환",
    )
    return parser.parse_args()


def main() -> None:
    base_dir = Path(__file__).resolve().parent
    args = parse_args()

    input_dir = resolve_dir(args.input_dir, base_dir)
    output_dir = resolve_dir(args.output_dir, base_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    input_dir.mkdir(parents=True, exist_ok=True)

    person_dirs = list_person_folders(input_dir)
    if not person_dirs:
        print(f"[WARN] No sub-folders found inside {input_dir}.")
        print("Add folders like 'haerin', 'no', ... and rerun.")
        return

    device = "cuda" if torch.cuda.is_available() else "cpu"
    print(f"[INFO] Using device: {device}")
    face_model = load_facenet_model(device)
    style_model = load_lpips_model(device, args.lpips_net)

    all_records: List[Dict[str, float]] = []
    for folder in person_dirs:
        try:
            person, records = process_folder(
                folder,
                face_model,
                style_model,
                device,
                args.style_threshold,
                args.style_distance_cap,
            )
            all_records.extend(records)
            print(
                f"[OK] {person.name}: '{person.original.name}' vs {len(records)} filtered images"
            )
        except Exception as exc:  # pylint: disable=broad-except
            print(f"[WARN] Failed on {folder.name}: {exc}")

    if not all_records:
        print("[WARN] No scores computed. Check folder layout and files.")
        return

    excel_path = output_dir / args.excel_name

    existing_order = load_existing_column_order(excel_path)
    existing_raw = read_existing_raw(excel_path)
    df_raw = merge_with_existing(all_records, existing_raw)
    if df_raw.empty:
        print("[WARN] Combined results are empty after merge.")
        return

    all_originals = list(dict.fromkeys(df_raw["original_name"].tolist()))
    column_order = existing_order + [name for name in all_originals if name not in existing_order]

    df_matrix = build_matrix(
        df_raw.to_dict(orient="records"),
        column_order if column_order else all_originals,
    )
    df_sorted = df_raw.sort_values(
        ["original_name", "final_percent"], ascending=[True, False]
    ).reset_index(drop=True)

    export_workbook(df_matrix, df_sorted, excel_path)
    cleanup_csv_artifacts(output_dir)

    print(f"[DONE] Excel saved to: {excel_path}")
    print(f"[INFO] Total rows: {len(df_sorted)}")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        sys.exit(1)
